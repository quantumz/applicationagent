"""
Job Application Tracker
Manages the applications spreadsheet from analysis results
"""

import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

PROJECT_ROOT = Path(__file__).parent.parent


class JobTracker:
    def __init__(self, spreadsheet_path):
        self.spreadsheet_path = Path(spreadsheet_path)
        self.workbook = None
        self.sheet = None

        self.headers = [
            'Job Title', 'Company', 'Location', 'Salary',
            'Fit Score', 'Decision', 'ATS Pass',
            'Applied?', 'Applied Date',
            'Response?', 'Response Date',
            'Interview?', 'Interview Date',
            'Final Status', 'Notes',
            'Job URL', 'Scraped Date'
        ]

    def load_or_create(self):
        """Load existing spreadsheet or create new one"""
        if self.spreadsheet_path.exists():
            print(f"📊 Loading existing spreadsheet: {self.spreadsheet_path}")
            self.workbook = load_workbook(self.spreadsheet_path)
            self.sheet = self.workbook.active
        else:
            print(f"📊 Creating new spreadsheet: {self.spreadsheet_path}")
            self.spreadsheet_path.parent.mkdir(parents=True, exist_ok=True)
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = "Job Applications"
            self._create_header()

    def _create_header(self):
        """Create formatted header row"""
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_num, header in enumerate(self.headers, 1):
            cell = self.sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        column_widths = {
            'A': 40, 'B': 25, 'C': 20, 'D': 20, 'E': 10, 'F': 15, 'G': 10,
            'H': 10, 'I': 15, 'J': 10, 'K': 15, 'L': 10, 'M': 15,
            'N': 15, 'O': 40, 'P': 50, 'Q': 15
        }
        for col_letter, width in column_widths.items():
            self.sheet.column_dimensions[col_letter].width = width

        self.sheet.freeze_panes = 'A2'

    def get_existing_keys(self):
        """Get set of dedup keys already in the spreadsheet (URL, or title|company for manual)"""
        if self.sheet.max_row == 1:
            return set()

        url_col = self.headers.index('Job URL') + 1
        title_col = self.headers.index('Job Title') + 1
        company_col = self.headers.index('Company') + 1
        keys = set()
        for row in range(2, self.sheet.max_row + 1):
            url = self.sheet.cell(row=row, column=url_col).value
            if url:
                keys.add(url)
            else:
                title = self.sheet.cell(row=row, column=title_col).value
                company = self.sheet.cell(row=row, column=company_col).value
                if title and company:
                    keys.add(f'{title}|{company}')
        return keys

    def add_jobs_from_db(self):
        """Add new jobs from DB, skipping rows already in the spreadsheet."""
        sys.path.insert(0, str(PROJECT_ROOT))
        from core.database import get_results
        all_results = get_results()
        existing_keys = self.get_existing_keys()
        new_jobs = 0

        for result in all_results:
            job = result['job_metadata']
            url = job.get('url', '') or ''
            dedup_key = url if url else f"{job.get('title', '')}|{job.get('company', '')}"
            if dedup_key and dedup_key in existing_keys:
                continue

            row_num = self.sheet.max_row + 1

            scraped_date = job.get('scraped_at', '')
            if scraped_date:
                try:
                    dt = datetime.fromisoformat(scraped_date.replace('Z', '+00:00'))
                    scraped_date = dt.strftime('%Y-%m-%d')
                except Exception:
                    pass

            try:
                fit_score = float(result.get('fit_score') or 0)
            except (TypeError, ValueError):
                fit_score = 0.0

            decision = result.get('decision', '')

            row_data = [
                job.get('title') or 'Unknown',
                job.get('company') or 'Unknown',
                job.get('location') or '',
                job.get('salary') or '',
                fit_score,
                decision,
                result.get('ai_analysis', {}).get('ats_pass_likelihood') or '',
                '',  # Applied?
                '',  # Applied Date
                '',  # Response?
                '',  # Response Date
                '',  # Interview?
                '',  # Interview Date
                'Pending',
                '',  # Notes
                url,
                scraped_date,
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = self.sheet.cell(row=row_num, column=col_num)
                cell.value = value
                if col_num in [5, 6, 7, 8, 10, 12]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                if col_num == 6:
                    if value == 'STRONG_MATCH':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(color="006100", bold=True)
                    elif value == 'APPLY':
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(color="9C6500")
                    elif value == 'MAYBE':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006")

            new_jobs += 1
            existing_keys.add(dedup_key)

        return new_jobs

    def save(self):
        self.workbook.save(self.spreadsheet_path)
        print(f"\n💾 Saved: {self.spreadsheet_path}")

    def print_summary(self):
        total_jobs = self.sheet.max_row - 1
        decision_col = self.headers.index('Decision') + 1
        decisions = {}
        for row in range(2, self.sheet.max_row + 1):
            decision = self.sheet.cell(row=row, column=decision_col).value
            decisions[decision] = decisions.get(decision, 0) + 1

        print(f"\n{'='*60}")
        print(f"SPREADSHEET SUMMARY — {total_jobs} jobs tracked")
        print(f"{'='*60}")
        for decision, count in sorted(decisions.items(), key=lambda x: x[1], reverse=True):
            print(f"  {decision}: {count}")
        print(f"{'='*60}\n")


def run_tracker(output_dir=None, data_dir=None):
    """
    Programmatic entry point for use from applicationagent.py.

    Args:
        output_dir: Directory for Excel output (default: output/excel/)
        data_dir: Unused — kept for API compatibility
    """
    output_dir = Path(output_dir) if output_dir else PROJECT_ROOT / 'output' / 'excel'
    spreadsheet_path = output_dir / 'job_tracker.xlsx'

    tracker = JobTracker(spreadsheet_path)
    tracker.load_or_create()
    new_jobs = tracker.add_jobs_from_db()
    tracker.save()
    tracker.print_summary()

    print(f"✅ Added {new_jobs} new job(s)")
    print(f"\nOpen: {spreadsheet_path}")


def main():
    import argparse

    parser = argparse.ArgumentParser(description='Job Application Tracker')
    parser.add_argument('--data-dir', default=None,
                        help='Directory containing analyzed JSON files (default: data/)')
    parser.add_argument('--output-dir', default=None,
                        help='Output directory for Excel spreadsheet (default: output/excel/)')

    args = parser.parse_args()

    print("="*60)
    print("JOB APPLICATION TRACKER")
    print("="*60)

    run_tracker(args.data_dir, args.output_dir)


if __name__ == "__main__":
    main()
