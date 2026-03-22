"""
Unit tests for scripts/tracker.py

Tests JobTracker and run_tracker without touching the real filesystem output
or the real database. tmp_path provides isolated Excel files; core.database
is mocked throughout.
"""

import pytest
from pathlib import Path
from unittest.mock import patch
from scripts.tracker import JobTracker, run_tracker


# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_result(title='Senior SRE', company='Acme Corp', location='Portland, OR',
                 salary='$150k', fit_score=0.92, decision='STRONG_MATCH',
                 url='https://example.com/job/1', scraped_at='2026-01-15T09:00:00',
                 ats='HIGH'):
    return {
        'fit_score': fit_score,
        'decision': decision,
        'ai_analysis': {'ats_pass_likelihood': ats},
        'job_metadata': {
            'title': title, 'company': company,
            'location': location, 'salary': salary,
            'url': url, 'scraped_at': scraped_at,
        },
    }


@pytest.fixture
def tracker(tmp_path):
    return JobTracker(tmp_path / 'test_tracker.xlsx')


@pytest.fixture
def fresh(tracker):
    """Tracker with a new spreadsheet loaded."""
    tracker.load_or_create()
    return tracker


# ── load_or_create ────────────────────────────────────────────────────────────

class TestLoadOrCreate:

    def test_creates_new_file_and_sets_sheet(self, tracker):
        tracker.load_or_create()
        assert tracker.workbook is not None
        assert tracker.sheet is not None

    def test_creates_header_row_on_new_file(self, tracker):
        tracker.load_or_create()
        assert tracker.sheet.cell(row=1, column=1).value == 'Job Title'
        assert tracker.sheet.cell(row=1, column=2).value == 'Company'

    def test_loads_existing_file(self, tmp_path):
        path = tmp_path / 'existing.xlsx'
        t1 = JobTracker(path)
        t1.load_or_create()
        t1.save()

        t2 = JobTracker(path)
        t2.load_or_create()
        # Existing header row must still be there
        assert t2.sheet.cell(row=1, column=1).value == 'Job Title'

    def test_freeze_panes_set(self, tracker):
        tracker.load_or_create()
        assert tracker.sheet.freeze_panes == 'A2'


# ── get_existing_keys ─────────────────────────────────────────────────────────

class TestGetExistingKeys:

    def test_empty_sheet_returns_empty_set(self, fresh):
        assert fresh.get_existing_keys() == set()

    def test_dedup_by_url(self, fresh):
        url_col = fresh.headers.index('Job URL') + 1
        fresh.sheet.cell(row=2, column=url_col).value = 'https://example.com/job/1'
        keys = fresh.get_existing_keys()
        assert 'https://example.com/job/1' in keys

    def test_dedup_by_title_company_when_no_url(self, fresh):
        title_col = fresh.headers.index('Job Title') + 1
        company_col = fresh.headers.index('Company') + 1
        fresh.sheet.cell(row=2, column=title_col).value = 'Senior SRE'
        fresh.sheet.cell(row=2, column=company_col).value = 'Acme Corp'
        keys = fresh.get_existing_keys()
        assert 'Senior SRE|Acme Corp' in keys

    def test_url_takes_priority_over_title_company(self, fresh):
        url_col = fresh.headers.index('Job URL') + 1
        title_col = fresh.headers.index('Job Title') + 1
        company_col = fresh.headers.index('Company') + 1
        fresh.sheet.cell(row=2, column=url_col).value = 'https://example.com/job/1'
        fresh.sheet.cell(row=2, column=title_col).value = 'Senior SRE'
        fresh.sheet.cell(row=2, column=company_col).value = 'Acme Corp'
        keys = fresh.get_existing_keys()
        assert 'https://example.com/job/1' in keys
        assert 'Senior SRE|Acme Corp' not in keys


# ── add_jobs_from_db ──────────────────────────────────────────────────────────

class TestAddJobsFromDb:

    def test_returns_count_of_added_jobs(self, fresh):
        results = [_make_result(url=f'https://example.com/job/{i}') for i in range(3)]
        with patch('core.database.get_results', return_value=results):
            count = fresh.add_jobs_from_db()
        assert count == 3

    def test_writes_title_and_company(self, fresh):
        with patch('core.database.get_results', return_value=[_make_result()]):
            fresh.add_jobs_from_db()
        title_col = fresh.headers.index('Job Title') + 1
        company_col = fresh.headers.index('Company') + 1
        assert fresh.sheet.cell(row=2, column=title_col).value == 'Senior SRE'
        assert fresh.sheet.cell(row=2, column=company_col).value == 'Acme Corp'

    def test_skips_duplicate_by_url(self, fresh):
        url = 'https://example.com/job/1'
        with patch('core.database.get_results', return_value=[_make_result(url=url)]):
            first = fresh.add_jobs_from_db()
        with patch('core.database.get_results', return_value=[_make_result(url=url)]):
            second = fresh.add_jobs_from_db()
        assert first == 1
        assert second == 0

    def test_skips_duplicate_by_title_company_when_no_url(self, fresh):
        result = _make_result(url='', title='Senior SRE', company='Acme Corp')
        with patch('core.database.get_results', return_value=[result]):
            first = fresh.add_jobs_from_db()
        with patch('core.database.get_results', return_value=[result]):
            second = fresh.add_jobs_from_db()
        assert first == 1
        assert second == 0

    def test_iso_date_formatted_to_ymd(self, fresh):
        with patch('core.database.get_results',
                   return_value=[_make_result(scraped_at='2026-01-15T09:00:00')]):
            fresh.add_jobs_from_db()
        date_col = fresh.headers.index('Scraped Date') + 1
        assert fresh.sheet.cell(row=2, column=date_col).value == '2026-01-15'

    def test_invalid_date_left_as_string(self, fresh):
        with patch('core.database.get_results',
                   return_value=[_make_result(scraped_at='not-a-date')]):
            fresh.add_jobs_from_db()
        date_col = fresh.headers.index('Scraped Date') + 1
        assert fresh.sheet.cell(row=2, column=date_col).value == 'not-a-date'

    def test_none_fit_score_coerced_to_zero(self, fresh):
        with patch('core.database.get_results',
                   return_value=[_make_result(fit_score=None)]):
            fresh.add_jobs_from_db()
        score_col = fresh.headers.index('Fit Score') + 1
        assert fresh.sheet.cell(row=2, column=score_col).value == 0.0

    def test_non_numeric_fit_score_coerced_to_zero(self, fresh):
        with patch('core.database.get_results',
                   return_value=[_make_result(fit_score='bad')]):
            fresh.add_jobs_from_db()
        score_col = fresh.headers.index('Fit Score') + 1
        assert fresh.sheet.cell(row=2, column=score_col).value == 0.0

    def test_empty_results_returns_zero(self, fresh):
        with patch('core.database.get_results', return_value=[]):
            count = fresh.add_jobs_from_db()
        assert count == 0

    def test_final_status_defaults_to_pending(self, fresh):
        with patch('core.database.get_results', return_value=[_make_result()]):
            fresh.add_jobs_from_db()
        status_col = fresh.headers.index('Final Status') + 1
        assert fresh.sheet.cell(row=2, column=status_col).value == 'Pending'


# ── print_summary ─────────────────────────────────────────────────────────────

class TestPrintSummary:

    def test_does_not_raise_on_empty_sheet(self, fresh, capsys):
        fresh.print_summary()
        out = capsys.readouterr().out
        assert '0 jobs tracked' in out

    def test_counts_decisions(self, fresh, capsys):
        results = [
            _make_result(decision='STRONG_MATCH', url='https://example.com/1'),
            _make_result(decision='STRONG_MATCH', url='https://example.com/2'),
            _make_result(decision='SKIP', url='https://example.com/3'),
        ]
        with patch('core.database.get_results', return_value=results):
            fresh.add_jobs_from_db()
        fresh.print_summary()
        out = capsys.readouterr().out
        assert 'STRONG_MATCH' in out
        assert 'SKIP' in out


# ── run_tracker ───────────────────────────────────────────────────────────────

class TestRunTracker:

    def test_creates_spreadsheet_file(self, tmp_path):
        output_dir = tmp_path / 'excel'
        with patch('core.database.get_results', return_value=[_make_result()]):
            run_tracker(output_dir=str(output_dir))
        assert (output_dir / 'job_tracker.xlsx').exists()

    def test_returns_without_error_on_empty_db(self, tmp_path):
        output_dir = tmp_path / 'excel'
        with patch('core.database.get_results', return_value=[]):
            run_tracker(output_dir=str(output_dir))
        assert (output_dir / 'job_tracker.xlsx').exists()

    def test_incremental_run_does_not_duplicate(self, tmp_path):
        output_dir = tmp_path / 'excel'
        result = _make_result()
        with patch('core.database.get_results', return_value=[result]):
            run_tracker(output_dir=str(output_dir))
        with patch('core.database.get_results', return_value=[result]):
            run_tracker(output_dir=str(output_dir))

        from openpyxl import load_workbook
        wb = load_workbook(output_dir / 'job_tracker.xlsx')
        # Only 1 data row despite two runs
        assert wb.active.max_row == 2
